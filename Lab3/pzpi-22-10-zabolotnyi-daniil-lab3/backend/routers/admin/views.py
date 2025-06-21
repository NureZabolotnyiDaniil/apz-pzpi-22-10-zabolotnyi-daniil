from datetime import timedelta, datetime
from typing import List
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import JSONResponse
import json
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.workbook import Workbook
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models.admins import Admin
from models.lanterns import Lantern
from models.parks import Park
from models.breakdowns import Breakdown
from models.renovations import Renovation
from models.repairmans import Repairman
from models.companies import Company
from models.db_activity import DatabaseActivity, ActivityType
from routers.admin.dependencies import get_current_admin, get_full_access_admin
from routers.admin.crud import (
    create_admin,
    authenticate_admin,
    create_access_token,
    get_all_admins,
    delete_admin,
    update_admin,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    update_admin_status,
)
from routers.admin.schemas import (
    RegistrationRequest,
    LoginRequest,
    AdminOut,
    AdminUpdate,
    AdminStatusUpdate,
)

router = APIRouter(prefix="/admin", tags=["admin"])


@router.post("/register")
async def register_admin(user: RegistrationRequest, db: Session = Depends(get_db)):
    try:
        create_admin(db, user)
        return JSONResponse(
            content={
                "message": "User registered successfully",
                "success": True
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
    except Exception as e:
        print(f"Registration error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={
                "detail": f"Registration error: {str(e)}",
                "success": False
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.post("/login")
async def login_admin(user: LoginRequest, db: Session = Depends(get_db)):
    try:
        admin = authenticate_admin(db, user.email, user.password)
        if not admin:
            return JSONResponse(
                status_code=401,
                content={
                    "detail": "Incorrect email or password",
                    "success": False
                },
                headers={
                    "WWW-Authenticate": "Bearer",
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                    "Access-Control-Allow-Headers": "*"
                }
            )
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": admin.email},
            expires_delta=access_token_expires,
        )
        return JSONResponse(
            content={
                "access_token": access_token, 
                "token_type": "bearer",
                "success": True
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
    except Exception as e:
        print(f"Login error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={
                "detail": f"Login error: {str(e)}",
                "success": False
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.get("/list", response_model=List[AdminOut])
async def get_admins_list(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    admins = get_all_admins(db)
    # Add park information to each admin
    admin_list = []
    for admin in admins:
        admin_dict = {
            "id": admin.id,
            "first_name": admin.first_name,
            "surname": admin.surname,
            "email": admin.email,
            "status": admin.status,
            "rights": admin.rights,
            "park_id": admin.park_id,
            "park_name": admin.park.name if admin.park else None
        }
        admin_list.append(admin_dict)
    return admin_list


@router.put("/edit", response_model=AdminOut)
async def set_admin(
    admin_data: AdminUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    updated_admin = update_admin(db, current_admin.id, admin_data)
    return updated_admin


@router.put("/update_status/{admin_email}", response_model=AdminOut)
async def set_admin_status(
    admin_email: str,
    status_update: AdminStatusUpdate,
    db: Session = Depends(get_db),
    full_access_admin: Admin = Depends(get_full_access_admin),
):
    updated_admin = update_admin_status(
        db, admin_email, status_update.status, status_update.rights
    )
    return updated_admin


@router.delete("/delete/{admin_id}", response_model=AdminOut)
async def remove_admin(
    admin_id: int,
    db: Session = Depends(get_db),
    full_access_admin: Admin = Depends(get_full_access_admin),
):
    deleted_admin = delete_admin(db, admin_id)
    return deleted_admin


@router.post("/export")
async def export_data(
    export_format: str = "json",
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Експорт даних системи в JSON або CSV форматі"""
    try:
        # Отримання всіх даних з кращою обробкою помилок
        try:
            lanterns = db.query(Lantern).all()
        except Exception as e:
            print(f"Error loading lanterns: {e}")
            lanterns = []
            
        try:
            parks = db.query(Park).all()
        except Exception as e:
            print(f"Error loading parks: {e}")
            parks = []
            
        try:
            breakdowns = db.query(Breakdown).all()
        except Exception as e:
            print(f"Error loading breakdowns: {e}")
            breakdowns = []
            
        try:
            renovations = db.query(Renovation).all()
        except Exception as e:
            print(f"Error loading renovations: {e}")
            renovations = []
            
        try:
            repairmen = db.query(Repairman).all()
        except Exception as e:
            print(f"Error loading repairmen: {e}")
            repairmen = []
            
        try:
            companies = db.query(Company).all()
        except Exception as e:
            print(f"Error loading companies: {e}")
            companies = []
        
        # Підготовка даних для експорту з кращою обробкою
        data = {
            "lanterns": [
                {
                    "id": l.id,
                    "name": getattr(l, 'name', f"Ліхтар #{l.id}"),
                    "park_id": getattr(l, 'park_id', None),
                    "brand": getattr(l, 'brand', None),
                    "model": getattr(l, 'model', None),
                    "power": getattr(l, 'power', None),
                    "height": getattr(l, 'height', None),
                    "status": getattr(l, 'status', 'unknown'),
                    "base_brightness": getattr(l, 'base_brightness', None),
                    "active_brightness": getattr(l, 'active_brightness', None),
                    "active_time": getattr(l, 'active_time', None),
                    "latitude": float(l.latitude) if getattr(l, 'latitude', None) else None,
                    "longitude": float(l.longitude) if getattr(l, 'longitude', None) else None,
                    "created_at": l.created_at.isoformat() if hasattr(l, 'created_at') and l.created_at else None,
                    "updated_at": l.updated_at.isoformat() if hasattr(l, 'updated_at') and l.updated_at else None
                } for l in lanterns
            ],
            "parks": [
                {
                    "id": p.id,
                    "name": getattr(p, 'name', 'Unnamed Park'),
                    "address": getattr(p, 'address', None),
                    "latitude": float(p.latitude) if getattr(p, 'latitude', None) else None,
                    "longitude": float(p.longitude) if getattr(p, 'longitude', None) else None,
                    "created_at": p.created_at.isoformat() if hasattr(p, 'created_at') and p.created_at else None,
                    "updated_at": p.updated_at.isoformat() if hasattr(p, 'updated_at') and p.updated_at else None
                } for p in parks
            ],
            "export_info": {
                "timestamp": datetime.utcnow().isoformat(),
                "exported_by": current_admin.email,
                "format": export_format,
                "total_records": len(lanterns) + len(parks) + len(breakdowns) + len(renovations) + len(repairmen) + len(companies)
            }
        }
        
        # Add other data types only if they exist and have records
        if breakdowns:
            data["breakdowns"] = [
                {
                    "id": b.id,
                    "lantern_id": getattr(b, 'lantern_id', None),
                    "description": getattr(b, 'description', ''),
                    "status": getattr(b, 'status', 'unknown'),
                    "priority": getattr(b, 'priority', None),
                    "reported_at": b.reported_at.isoformat() if hasattr(b, 'reported_at') and b.reported_at else None,
                    "fixed_at": b.fixed_at.isoformat() if hasattr(b, 'fixed_at') and b.fixed_at else None
                } for b in breakdowns
            ]
            
        if renovations:
            data["renovations"] = [
                {
                    "id": r.id,
                    "lantern_id": getattr(r, 'lantern_id', None),
                    "description": getattr(r, 'description', ''),
                    "status": getattr(r, 'status', 'unknown'),
                    "priority": getattr(r, 'priority', None),
                    "cost": getattr(r, 'cost', None),
                    "repairman_id": getattr(r, 'repairman_id', None)
                } for r in renovations
            ]
            
        if repairmen:
            data["repairmen"] = [
                {
                    "id": r.id,
                    "first_name": getattr(r, 'first_name', 'Unknown'),
                    "surname": getattr(r, 'surname', ''),
                    "phone": getattr(r, 'phone', None),
                    "email": getattr(r, 'email', None),
                    "company_id": getattr(r, 'company_id', None)
                } for r in repairmen
            ]
            
        if companies:
            data["companies"] = [
                {
                    "id": c.id,
                    "name": getattr(c, 'name', 'Unnamed Company'),
                    "address": getattr(c, 'address', None),
                    "phone": getattr(c, 'phone', None),
                    "email": getattr(c, 'email', None)
                } for c in companies
            ]
        
        # Логування активності (опціонально)
        try:
            activity = DatabaseActivity(
                activity_type=ActivityType.DATA_EXPORT,
                entity_type="system",
                description=f"Експорт даних у форматі {export_format.upper()}",
                details=json.dumps({"format": export_format, "records_count": data["export_info"]["total_records"]}),
                performed_by=current_admin.email
            )
            db.add(activity)
            db.commit()
        except Exception as log_error:
            print(f"Warning: Could not log export activity: {log_error}")
        
        # Повернення результату
        return JSONResponse(
            content={"data": data, "format": export_format, "success": True},
            headers={
                "Content-Disposition": f"attachment; filename=smartlighting_export.{export_format}",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
            
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Помилка експорту даних: {str(e)}",
                "success": False,
                "details": str(e)
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.post("/export/excel")
async def export_excel(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Експорт даних системи в Excel форматі"""
    try:
        # Отримання всіх даних
        try:
            lanterns = db.query(Lantern).all()
        except Exception as e:
            print(f"Error loading lanterns: {e}")
            lanterns = []
            
        try:
            parks = db.query(Park).all()
        except Exception as e:
            print(f"Error loading parks: {e}")
            parks = []
            
        try:
            breakdowns = db.query(Breakdown).all()
        except Exception as e:
            print(f"Error loading breakdowns: {e}")
            breakdowns = []
            
        try:
            renovations = db.query(Renovation).all()
        except Exception as e:
            print(f"Error loading renovations: {e}")
            renovations = []
            
        try:
            repairmen = db.query(Repairman).all()
        except Exception as e:
            print(f"Error loading repairmen: {e}")
            repairmen = []
            
        try:
            companies = db.query(Company).all()
        except Exception as e:
            print(f"Error loading companies: {e}")
            companies = []

        # Створення Excel файлу
        workbook = Workbook()
        
        # Видалити стандартний лист
        workbook.remove(workbook.active)
        
        # Лист "Ліхтарі"
        if lanterns:
            ws_lanterns = workbook.create_sheet(title="Ліхтарі")
            # Заголовки
            headers = ["ID", "Назва", "Парк ID", "Бренд", "Модель", "Потужність", "Висота", "Статус", 
                      "Базова яскравість", "Активна яскравість", "Час активності", "Широта", "Довгота", 
                      "Створено", "Оновлено"]
            ws_lanterns.append(headers)
            
            # Дані
            for l in lanterns:
                row = [
                    l.id,
                    getattr(l, 'name', f"Ліхтар #{l.id}"),
                    getattr(l, 'park_id', None),
                    getattr(l, 'brand', None),
                    getattr(l, 'model', None),
                    getattr(l, 'power', None),
                    getattr(l, 'height', None),
                    getattr(l, 'status', 'unknown'),
                    getattr(l, 'base_brightness', None),
                    getattr(l, 'active_brightness', None),
                    getattr(l, 'active_time', None),
                    float(l.latitude) if getattr(l, 'latitude', None) else None,
                    float(l.longitude) if getattr(l, 'longitude', None) else None,
                    l.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(l, 'created_at') and l.created_at else None,
                    l.updated_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(l, 'updated_at') and l.updated_at else None
                ]
                ws_lanterns.append(row)
        
        # Лист "Парки"
        if parks:
            ws_parks = workbook.create_sheet(title="Парки")
            headers = ["ID", "Назва", "Адреса", "Широта", "Довгота", "Створено", "Оновлено"]
            ws_parks.append(headers)
            
            for p in parks:
                row = [
                    p.id,
                    getattr(p, 'name', 'Unnamed Park'),
                    getattr(p, 'address', None),
                    float(p.latitude) if getattr(p, 'latitude', None) else None,
                    float(p.longitude) if getattr(p, 'longitude', None) else None,
                    p.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p, 'created_at') and p.created_at else None,
                    p.updated_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(p, 'updated_at') and p.updated_at else None
                ]
                ws_parks.append(row)
        
        # Лист "Поломки"
        if breakdowns:
            ws_breakdowns = workbook.create_sheet(title="Поломки")
            headers = ["ID", "Ліхтар ID", "Опис", "Статус", "Пріоритет", "Повідомлено", "Виправлено"]
            ws_breakdowns.append(headers)
            
            for b in breakdowns:
                row = [
                    b.id,
                    getattr(b, 'lantern_id', None),
                    getattr(b, 'description', ''),
                    getattr(b, 'status', 'unknown'),
                    getattr(b, 'priority', None),
                    b.reported_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(b, 'reported_at') and b.reported_at else None,
                    b.fixed_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(b, 'fixed_at') and b.fixed_at else None
                ]
                ws_breakdowns.append(row)
        
        # Лист "Ремонти"
        if renovations:
            ws_renovations = workbook.create_sheet(title="Ремонти")
            headers = ["ID", "Ліхтар ID", "Опис", "Статус", "Дата початку", "Дата закінчення"]
            ws_renovations.append(headers)
            
            for r in renovations:
                row = [
                    r.id,
                    getattr(r, 'lantern_id', None),
                    getattr(r, 'description', ''),
                    getattr(r, 'status', 'unknown'),
                    r.start_date.strftime("%Y-%m-%d %H:%M:%S") if getattr(r, 'start_date', None) else None,
                    r.end_date.strftime("%Y-%m-%d %H:%M:%S") if getattr(r, 'end_date', None) else None
                ]
                ws_renovations.append(row)
        
        # Лист "Ремонтники"
        if repairmen:
            ws_repairmen = workbook.create_sheet(title="Ремонтники")
            headers = ["ID", "Ім'я", "Прізвище", "Телефон", "Email", "Компанія ID"]
            ws_repairmen.append(headers)
            
            for r in repairmen:
                row = [
                    r.id,
                    getattr(r, 'first_name', 'Unknown'),
                    getattr(r, 'surname', ''),
                    getattr(r, 'phone', None),
                    getattr(r, 'email', None),
                    getattr(r, 'company_id', None)
                ]
                ws_repairmen.append(row)
        
        # Лист "Компанії"
        if companies:
            ws_companies = workbook.create_sheet(title="Компанії")
            headers = ["ID", "Назва", "Адреса", "Телефон", "Email"]
            ws_companies.append(headers)
            
            for c in companies:
                row = [
                    c.id,
                    getattr(c, 'name', 'Unnamed Company'),
                    getattr(c, 'address', None),
                    getattr(c, 'phone', None),
                    getattr(c, 'email', None)
                ]
                ws_companies.append(row)
        
        # Лист "Інформація"
        ws_info = workbook.create_sheet(title="Інформація")
        ws_info.append(["Параметр", "Значення"])
        ws_info.append(["Дата експорту", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
        ws_info.append(["Експортував", current_admin.email])
        ws_info.append(["Кількість ліхтарів", len(lanterns)])
        ws_info.append(["Кількість парків", len(parks)])
        ws_info.append(["Кількість поломок", len(breakdowns)])
        ws_info.append(["Кількість ремонтів", len(renovations)])
        ws_info.append(["Кількість ремонтників", len(repairmen)])
        ws_info.append(["Кількість компаній", len(companies)])
        
        # Якщо немає даних - створити порожній лист
        if not workbook.worksheets:
            ws_empty = workbook.create_sheet(title="Пусто")
            ws_empty.append(["Немає даних для експорту"])
        
        # Збереження в пам'ять
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Логування активності
        try:
            activity = DatabaseActivity(
                activity_type=ActivityType.DATA_EXPORT,
                entity_type="system",
                description="Експорт даних у форматі Excel",
                details=json.dumps({"format": "excel", "records_count": len(lanterns) + len(parks) + len(breakdowns) + len(renovations) + len(repairmen) + len(companies)}),
                performed_by=current_admin.email
            )
            db.add(activity)
            db.commit()
        except Exception as log_error:
            print(f"Warning: Could not log export activity: {log_error}")
        
        # Повернення файлу
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"smartlighting_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            BytesIO(excel_file.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
            
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Помилка експорту Excel: {str(e)}",
                "success": False,
                "details": str(e)
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.post("/backup")
async def create_backup(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Створення повного бекапу системи"""
    try:
        # Отримання всіх даних з усіх таблиць
        admins = db.query(Admin).all()
        lanterns = db.query(Lantern).all()
        parks = db.query(Park).all()
        breakdowns = db.query(Breakdown).all()
        renovations = db.query(Renovation).all()
        repairmen = db.query(Repairman).all()
        companies = db.query(Company).all()
        activities = db.query(DatabaseActivity).all()
        
        backup_data = {
            "admins": [
                {
                    "id": a.id,
                    "email": a.email,
                    "first_name": a.first_name,
                    "surname": a.surname,
                    "status": a.status,
                    "rights": a.rights,
                    "park_id": getattr(a, 'park_id', None),
                    "created_at": a.created_at.isoformat() if hasattr(a, 'created_at') and a.created_at else None
                } for a in admins
            ],
            "lanterns": [
                {
                    "id": l.id,
                    "name": getattr(l, 'name', f"Ліхтар #{l.id}"),
                    "park_id": l.park_id,
                    "brand": getattr(l, 'brand', None),
                    "model": getattr(l, 'model', None),
                    "power": getattr(l, 'power', None),
                    "height": getattr(l, 'height', None),
                    "status": l.status,
                    "base_brightness": getattr(l, 'base_brightness', None),
                    "active_brightness": getattr(l, 'active_brightness', None),
                    "active_time": getattr(l, 'active_time', None),
                    "latitude": float(l.latitude) if l.latitude else None,
                    "longitude": float(l.longitude) if l.longitude else None,
                    "created_at": l.created_at.isoformat() if hasattr(l, 'created_at') and l.created_at else None,
                    "updated_at": l.updated_at.isoformat() if hasattr(l, 'updated_at') and l.updated_at else None
                } for l in lanterns
            ],
            "parks": [
                {
                    "id": p.id,
                    "name": p.name,
                    "address": getattr(p, 'address', None),
                    "area": getattr(p, 'area', None),
                    "latitude": float(p.latitude) if p.latitude else None,
                    "longitude": float(p.longitude) if p.longitude else None,
                    "created_at": p.created_at.isoformat() if hasattr(p, 'created_at') and p.created_at else None,
                    "updated_at": p.updated_at.isoformat() if hasattr(p, 'updated_at') and p.updated_at else None
                } for p in parks
            ],
            "breakdowns": [
                {
                    "id": b.id,
                    "lantern_id": b.lantern_id,
                    "description": b.description,
                    "status": b.status,
                    "priority": getattr(b, 'priority', None),
                    "reported_at": b.reported_at.isoformat() if hasattr(b, 'reported_at') and b.reported_at else None,
                    "fixed_at": b.fixed_at.isoformat() if hasattr(b, 'fixed_at') and b.fixed_at else None
                } for b in breakdowns
            ],
            "renovations": [
                {
                    "id": r.id,
                    "lantern_id": r.lantern_id,
                    "description": getattr(r, 'description', None),
                    "status": r.status,
                    "priority": getattr(r, 'priority', None),
                    "cost": getattr(r, 'cost', None),
                    "repairman_id": getattr(r, 'repairman_id', None)
                } for r in renovations
            ],
            "repairmen": [
                {
                    "id": r.id,
                    "first_name": r.first_name,
                    "last_name": getattr(r, 'last_name', None),
                    "surname": getattr(r, 'surname', None),
                    "phone": getattr(r, 'phone', None),
                    "email": getattr(r, 'email', None),
                    "specialization": getattr(r, 'specialization', None),
                    "company_id": getattr(r, 'company_id', None)
                } for r in repairmen
            ],
            "companies": [
                {
                    "id": c.id,
                    "name": c.name,
                    "address": getattr(c, 'address', None),
                    "phone": getattr(c, 'phone', None),
                    "email": getattr(c, 'email', None),
                    "notes": getattr(c, 'notes', None)
                } for c in companies
            ],
            "activities": [
                {
                    "id": a.id,
                    "activity_type": a.activity_type.value if a.activity_type else None,
                    "entity_type": a.entity_type,
                    "entity_id": a.entity_id,
                    "description": a.description,
                    "details": a.details,
                    "performed_by": a.performed_by,
                    "created_at": a.created_at.isoformat() if a.created_at else None
                } for a in activities
            ],
            "backup_info": {
                "timestamp": str(datetime.utcnow()),
                "created_by": current_admin.email,
                "version": "1.0",
                "total_records": len(admins) + len(lanterns) + len(parks) + len(breakdowns) + len(renovations) + len(repairmen) + len(companies)
            }
        }
        
        # Логування створення бекапу
        try:
            activity = DatabaseActivity(
                activity_type=ActivityType.BACKUP_CREATED,
                entity_type="system",
                description="Створено повний бекап системи",
                details=json.dumps({"total_records": backup_data["backup_info"]["total_records"]}),
                performed_by=current_admin.email
            )
            db.add(activity)
            db.commit()
        except Exception as log_error:
            print(f"Warning: Could not log backup activity: {log_error}")
        
        # Generate backup filename with timestamp
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"smartlighting_backup_{timestamp}.json"
        
        return JSONResponse(
            content={
                "data": backup_data,
                "name": f"smartlighting_backup_{timestamp}",
                "filename": filename,
                "success": True
            },
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
        
    except Exception as e:
        print(f"Backup error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Помилка створення бекапу: {str(e)}",
                "success": False,
                "details": str(e)
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Імпорт даних з JSON файлу"""
    try:
        # Read uploaded file content
        file_content = await file.read()
        if not file_content:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "Файл не надано або порожній",
                    "success": False
                },
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                    "Access-Control-Allow-Headers": "*"
                }
            )
        
        # Parse JSON data
        try:
            import_data = json.loads(file_content.decode('utf-8'))
        except json.JSONDecodeError as e:
            return JSONResponse(
                status_code=400,
                content={
                    "error": f"Невірний формат JSON: {str(e)}",
                    "success": False
                },
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                    "Access-Control-Allow-Headers": "*"
                }
            )
        
        imported_counts = {
            "lanterns": 0,
            "parks": 0,
            "breakdowns": 0,
            "renovations": 0,
            "repairmen": 0,
            "companies": 0
        }
        
        # Import data (this is a basic implementation - in production you'd want more sophisticated merging)
        if "data" in import_data:
            data = import_data["data"]
        else:
            data = import_data
            
        # Import parks first (as they are referenced by other entities)
        if "parks" in data and isinstance(data["parks"], list):
            for park_data in data["parks"]:
                try:
                    # Check if park already exists
                    existing_park = db.query(Park).filter(Park.id == park_data.get("id")).first()
                    if not existing_park:
                        park = Park(
                            id=park_data.get("id"),
                            name=park_data.get("name", "Imported Park"),
                            address=park_data.get("address"),
                            area=park_data.get("area"),
                            latitude=park_data.get("latitude"),
                            longitude=park_data.get("longitude")
                        )
                        db.add(park)
                        imported_counts["parks"] += 1
                except Exception as e:
                    print(f"Error importing park {park_data.get('id', 'unknown')}: {e}")
                    
        # Import companies
        if "companies" in data and isinstance(data["companies"], list):
            for company_data in data["companies"]:
                try:
                    existing_company = db.query(Company).filter(Company.id == company_data.get("id")).first()
                    if not existing_company:
                        company = Company(
                            id=company_data.get("id"),
                            name=company_data.get("name", "Imported Company"),
                            address=company_data.get("address"),
                            phone=company_data.get("phone"),
                            email=company_data.get("email"),
                            notes=company_data.get("notes")
                        )
                        db.add(company)
                        imported_counts["companies"] += 1
                except Exception as e:
                    print(f"Error importing company {company_data.get('id', 'unknown')}: {e}")
                    
        # Import repairmen
        if "repairmen" in data and isinstance(data["repairmen"], list):
            for repairman_data in data["repairmen"]:
                try:
                    existing_repairman = db.query(Repairman).filter(Repairman.id == repairman_data.get("id")).first()
                    if not existing_repairman:
                        repairman = Repairman(
                            id=repairman_data.get("id"),
                            first_name=repairman_data.get("first_name", "Unknown"),
                            surname=repairman_data.get("surname", ""),
                            phone=repairman_data.get("phone"),
                            email=repairman_data.get("email"),
                            company_id=repairman_data.get("company_id")
                        )
                        db.add(repairman)
                        imported_counts["repairmen"] += 1
                except Exception as e:
                    print(f"Error importing repairman {repairman_data.get('id', 'unknown')}: {e}")
                    
        # Import lanterns
        if "lanterns" in data and isinstance(data["lanterns"], list):
            for lantern_data in data["lanterns"]:
                try:
                    existing_lantern = db.query(Lantern).filter(Lantern.id == lantern_data.get("id")).first()
                    if not existing_lantern:
                        lantern = Lantern(
                            id=lantern_data.get("id"),
                            name=lantern_data.get("name", f"Ліхтар #{lantern_data.get('id')}"),
                            park_id=lantern_data.get("park_id"),
                            brand=lantern_data.get("brand"),
                            model=lantern_data.get("model"),
                            power=lantern_data.get("power"),
                            height=lantern_data.get("height"),
                            status=lantern_data.get("status", "working"),
                            base_brightness=lantern_data.get("base_brightness"),
                            active_brightness=lantern_data.get("active_brightness"),
                            active_time=lantern_data.get("active_time"),
                            latitude=lantern_data.get("latitude"),
                            longitude=lantern_data.get("longitude")
                        )
                        db.add(lantern)
                        imported_counts["lanterns"] += 1
                except Exception as e:
                    print(f"Error importing lantern {lantern_data.get('id', 'unknown')}: {e}")
                    
        # Import breakdowns
        if "breakdowns" in data and isinstance(data["breakdowns"], list):
            for breakdown_data in data["breakdowns"]:
                try:
                    existing_breakdown = db.query(Breakdown).filter(Breakdown.id == breakdown_data.get("id")).first()
                    if not existing_breakdown:
                        breakdown = Breakdown(
                            id=breakdown_data.get("id"),
                            lantern_id=breakdown_data.get("lantern_id"),
                            description=breakdown_data.get("description", ""),
                            status=breakdown_data.get("status", "reported"),
                            priority=breakdown_data.get("priority", "medium")
                        )
                        db.add(breakdown)
                        imported_counts["breakdowns"] += 1
                except Exception as e:
                    print(f"Error importing breakdown {breakdown_data.get('id', 'unknown')}: {e}")
                    
        # Import renovations
        if "renovations" in data and isinstance(data["renovations"], list):
            for renovation_data in data["renovations"]:
                try:
                    existing_renovation = db.query(Renovation).filter(Renovation.id == renovation_data.get("id")).first()
                    if not existing_renovation:
                        renovation = Renovation(
                            id=renovation_data.get("id"),
                            lantern_id=renovation_data.get("lantern_id"),
                            description=renovation_data.get("description", ""),
                            status=renovation_data.get("status", "planned"),
                            priority=renovation_data.get("priority", "medium"),
                            cost=renovation_data.get("cost", 0),
                            repairman_id=renovation_data.get("repairman_id")
                        )
                        db.add(renovation)
                        imported_counts["renovations"] += 1
                except Exception as e:
                    print(f"Error importing renovation {renovation_data.get('id', 'unknown')}: {e}")
        
        # Commit all changes
        db.commit()
            
        # Log import activity
        try:
            activity = DatabaseActivity(
                activity_type=ActivityType.DATA_EXPORT,  # Using DATA_EXPORT as closest match
                entity_type="system",
                description="Імпорт даних з файлу",
                details=json.dumps({"imported_counts": imported_counts}),
                performed_by=current_admin.email
            )
            db.add(activity)
            db.commit()
        except Exception as log_error:
            print(f"Warning: Could not log import activity: {log_error}")
        
        return JSONResponse(
            content={
                "message": "Дані успішно імпортовано",
                "imported_counts": imported_counts,
                "success": True
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
        
    except Exception as e:
        print(f"Import error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Помилка імпорту даних: {str(e)}",
                "success": False,
                "details": str(e)
            },
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.get("/export-test")
async def export_test(
    db: Session = Depends(get_db),
):
    """Тестовий експорт без авторизації"""
    try:
        lanterns = db.query(Lantern).all()
        parks = db.query(Park).all()
        
        data = {
            "lanterns_count": len(lanterns),
            "parks_count": len(parks),
            "test": True,
            "timestamp": str(datetime.utcnow())
        }
        
        return JSONResponse(
            content={"data": data, "success": True},
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": str(e), "success": False},
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
                "Access-Control-Allow-Headers": "*"
            }
        )


@router.get("/me", response_model=AdminOut)
async def get_current_user(
    current_admin: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    return {
        "id": current_admin.id,
        "first_name": current_admin.first_name,
        "surname": current_admin.surname,
        "email": current_admin.email,
        "status": current_admin.status,
        "rights": current_admin.rights,
        "park_id": current_admin.park_id,
        "park_name": current_admin.park.name if current_admin.park else None
    }
